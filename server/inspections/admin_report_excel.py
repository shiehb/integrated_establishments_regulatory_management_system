"""
Admin Report Excel Generator using openpyxl
Generates professional Excel reports with DENR official standards
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class AdminReportExcelGenerator:
    """
    Professional Excel generator for admin reports (establishments and users)
    """
    
    def __init__(self, report_data, filters_applied):
        self.report_data = report_data
        self.filters_applied = filters_applied
        self.workbook = Workbook()
        
        # Generate DENR reference number
        self.reference_number = self._generate_reference_number()
        
        # DENR colors
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')  # DENR Blue
        self.denr_green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # DENR Green
        self.subheader_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        self.light_blue_fill = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')
        self.light_green_fill = PatternFill(start_color='E7F7E7', end_color='E7F7E7', fill_type='solid')
        self.light_red_fill = PatternFill(start_color='FFE7E7', end_color='FFE7E7', fill_type='solid')
        
        # Define fonts - Arial as per DENR standards
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        # Define borders
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def _generate_reference_number(self):
        """Generate DENR reference number: EIA-YYYY-MM-DD-####"""
        now = datetime.now()
        date_str = now.strftime('%Y-%m-%d')
        sequence = str(int(now.timestamp() * 1000))[-4:]
        return f"EIA-{date_str}-{sequence}"
    
    def _add_denr_header(self, worksheet, start_row=1):
        """Add official DENR header to worksheet"""
        ws = worksheet
        row = start_row
        
        # DENR Header
        ws[f'A{row}'] = 'REPUBLIC OF THE PHILIPPINES'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='0066CC')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        ws[f'A{row}'] = 'DEPARTMENT OF ENVIRONMENT AND NATURAL RESOURCES'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='008000')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        ws[f'A{row}'] = 'ENVIRONMENTAL MANAGEMENT BUREAU'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='008000')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        ws[f'A{row}'] = 'REGION I'
        ws[f'A{row}'].font = Font(name='Arial', size=10, color='008000')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        ws[f'A{row}'] = 'Kalikasang Protektado, Paglilingkod na Tapat.'
        ws[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color='666666')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        ws[f'A{row}'] = f'Reference Number: {self.reference_number}'
        ws[f'A{row}'].font = self.bold_font
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        return row
    
    def _freeze_header_row(self, worksheet, header_row):
        """Freeze panes at header row"""
        worksheet.freeze_panes = worksheet[f'A{header_row + 1}']
    
    def _add_auto_filter(self, worksheet, start_row, end_row, num_cols):
        """Add auto-filter to data table"""
        start_col_letter = get_column_letter(1)
        end_col_letter = get_column_letter(num_cols)
        worksheet.auto_filter.ref = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        
    def _auto_adjust_column_width(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def generate_establishments_report(self):
        """Generate establishments Excel report"""
        ws = self.workbook.active
        ws.title = 'Establishments'
        
        # Add DENR header
        row = self._add_denr_header(ws)
        
        # Title
        ws[f'A{row}'] = 'ADMIN REPORT - ESTABLISHMENTS'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Generation info
        ws[f'A{row}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws[f'A{row}'].font = self.normal_font
        row += 1
        
        # Filters applied
        row = 4
        ws[f'A{row}'] = 'FILTERS APPLIED'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        for key, value in self.filters_applied.items():
            if value and str(value) != 'ALL':
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'A{row}'].font = self.bold_font
                ws[f'B{row}'] = str(value)
                row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = 'SUMMARY STATISTICS'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        total = len(self.report_data)
        active = sum(1 for r in self.report_data if r.get('is_active', False))
        inactive = total - active
        
        row += 1
        stats_data = [
            ['Total Establishments', total],
            ['Active', active],
            ['Inactive', inactive],
        ]
        
        for stat in stats_data:
            ws[f'A{row}'] = stat[0]
            ws[f'A{row}'].font = self.bold_font
            ws[f'B{row}'] = stat[1]
            row += 1
        
        # Data table
        row += 2
        header_row = row
        headers = ['Name', 'Nature of Business', 'Province', 'City', 'Barangay', 'Date Added', 'Status']
        
        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        row += 1
        data_start_row = row
        
        # Data rows
        for record in self.report_data:
            created_at = record.get('created_at', '')
            if created_at:
                try:
                    if 'T' in created_at:
                        created_at = created_at.split('T')[0]
                    created_at = created_at[:10]
                except:
                    pass
            
            status = 'Active' if record.get('is_active', False) else 'Inactive'
            
            row_data = [
                record.get('name', 'N/A'),
                record.get('nature_of_business', 'N/A'),
                record.get('province', 'N/A'),
                record.get('city', 'N/A'),
                record.get('barangay', 'N/A'),
                created_at or 'N/A',
                status,
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Apply status color
                if col == 7:  # Status column
                    if record.get('is_active', False):
                        cell.fill = self.light_green_fill
                    else:
                        cell.fill = self.light_red_fill
                elif row % 2 == 0:  # Alternate row color
                    cell.fill = self.light_blue_fill
            
            row += 1
        
        self._auto_adjust_column_width(ws)
    
    def _add_routing_section(self, worksheet, start_row):
        """Add DENR routing section for workflow tracking"""
        ws = worksheet
        row = start_row + 2
        
        ws[f'A{row}'] = 'ROUTING AND APPROVAL'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        routing_headers = ['Stage', 'Name', 'Position', 'Date', 'Signature']
        for col, header in enumerate(routing_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        row += 1
        routing_data = [
            ['Prepared by', '', 'Admin Staff', '', ''],
            ['Reviewed by', '', 'Section Chief', '', ''],
            ['Recommended by', '', 'Division Chief', '', ''],
            ['Approved by', '', 'Regional Director', '', ''],
        ]
        
        for routing_row in routing_data:
            for col, value in enumerate(routing_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row % 2 == 0:
                    cell.fill = self.light_blue_fill
            row += 1
    
    def generate_users_report(self):
        """Generate users Excel report"""
        ws = self.workbook.active
        ws.title = 'Users'
        
        # Add DENR header
        row = self._add_denr_header(ws)
        
        # Title
        ws[f'A{row}'] = 'ADMIN REPORT - USERS'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Generation info
        ws[f'A{row}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws[f'A{row}'].font = self.normal_font
        row += 1
        
        # Filters applied
        row = 4
        ws[f'A{row}'] = 'FILTERS APPLIED'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        for key, value in self.filters_applied.items():
            if value and str(value) != 'ALL':
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'A{row}'].font = self.bold_font
                ws[f'B{row}'] = str(value)
                row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = 'SUMMARY STATISTICS'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        total = len(self.report_data)
        active = sum(1 for r in self.report_data if r.get('is_active', False))
        inactive = total - active
        
        row += 1
        stats_data = [
            ['Total Users', total],
            ['Active', active],
            ['Inactive', inactive],
        ]
        
        for stat in stats_data:
            ws[f'A{row}'] = stat[0]
            ws[f'A{row}'].font = self.bold_font
            ws[f'B{row}'] = stat[1]
            row += 1
        
        # Data table
        row += 2
        header_row = row
        headers = ['Name', 'Email', 'User Level', 'Section', 'Date Joined', 'Last Updated', 'Status']
        
        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        row += 1
        data_start_row = row
        
        # Data rows
        for record in self.report_data:
            date_joined = record.get('date_joined', '')
            if date_joined:
                try:
                    if 'T' in date_joined:
                        date_joined = date_joined.split('T')[0]
                    date_joined = date_joined[:10]
                except:
                    pass
            
            updated_at = record.get('updated_at', '')
            if updated_at:
                try:
                    if 'T' in updated_at:
                        updated_at = updated_at.split('T')[0]
                    updated_at = updated_at[:10]
                except:
                    pass
            
            status = 'Active' if record.get('is_active', False) else 'Inactive'
            
            row_data = [
                record.get('full_name', record.get('email', 'N/A')),
                record.get('email', 'N/A'),
                record.get('userlevel', 'N/A'),
                record.get('section', 'N/A') or 'N/A',
                date_joined or 'N/A',
                updated_at or 'N/A',
                status,
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Apply status color
                if col == 7:  # Status column
                    if record.get('is_active', False):
                        cell.fill = self.light_green_fill
                    else:
                        cell.fill = self.light_red_fill
                elif row % 2 == 0:  # Alternate row color
                    cell.fill = self.light_blue_fill
            
            row += 1
        
        self._auto_adjust_column_width(ws)

