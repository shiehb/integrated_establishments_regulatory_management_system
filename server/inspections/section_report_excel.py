"""
Division Report Excel Generator using openpyxl
Generates professional Excel reports with multiple worksheets
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DivisionReportExcelGenerator:
    """
    Professional Excel generator for division reports
    """
    
    def __init__(self, report_data, filters_applied):
        self.report_data = report_data
        self.filters_applied = filters_applied
        self.workbook = Workbook()
        
        # Define colors
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        self.light_blue_fill = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')
        self.light_green_fill = PatternFill(start_color='E7F7E7', end_color='E7F7E7', fill_type='solid')
        self.light_red_fill = PatternFill(start_color='FFE7E7', end_color='FFE7E7', fill_type='solid')
        
        # Define fonts
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        # Define borders
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
    def _auto_adjust_column_width(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self):
        """Create summary statistics worksheet"""
        ws = self.workbook.active
        ws.title = 'Summary Statistics'
        
        # Title
        ws['A1'] = 'DIVISION REPORT - SUMMARY STATISTICS'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        
        # Generation info
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = self.normal_font
        
        # Filters applied
        row = 4
        ws[f'A{row}'] = 'FILTERS APPLIED'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for key, value in self.filters_applied.items():
            if value:
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = self.bold_font
                row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = 'INSPECTION SUMMARY'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.light_blue_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        stats = self.report_data.get('statistics', {})
        inspection_stats = stats.get('inspection_summary', {})
        compliance_stats = stats.get('compliance_summary', {})
        
        row += 1
        summary_data = [
            ('Total Inspections', inspection_stats.get('total_inspections', 0)),
            ('Division Reviewed', inspection_stats.get('division_reviewed', 0)),
            ('Section Completed', inspection_stats.get('section_completed', 0)),
            ('Total NOV Issued', inspection_stats.get('total_nov', 0)),
            ('Total NOO Issued', inspection_stats.get('total_noo', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.bold_font
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Compliance summary
        row += 2
        ws[f'A{row}'] = 'COMPLIANCE SUMMARY'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.light_green_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        compliance_data = [
            ('Compliant', compliance_stats.get('compliant_count', 0)),
            ('Non-Compliant', compliance_stats.get('non_compliant_count', 0)),
            ('Pending', compliance_stats.get('pending_count', 0)),
        ]
        
        for label, value in compliance_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.bold_font
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        self._auto_adjust_column_width(ws)
    
    def _create_detailed_data_sheet(self):
        """Create detailed data worksheet"""
        ws = self.workbook.create_sheet(title='Detailed Data')
        
        # Headers
        headers = [
            'Inspection No.', 'Establishment', 'Law', 'Inspection Date',
            'Status', 'NOV', 'NOO', 'Compliance Status', 'Inspected By'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        records = self.report_data.get('records', [])
        for row_idx, record in enumerate(records, start=2):
            # Inspection No.
            ws.cell(row=row_idx, column=1, value=record.get('code', 'N/A')).border = self.border
            
            # Establishment
            ws.cell(row=row_idx, column=2, value=record.get('establishment_name', 'N/A')).border = self.border
            
            # Law
            ws.cell(row=row_idx, column=3, value=record.get('law', 'N/A')).border = self.border
            
            # Inspection Date
            created_at = record.get('created_at', '')
            if created_at:
                created_at = created_at[:10]  # Extract date part
            ws.cell(row=row_idx, column=4, value=created_at or 'N/A').border = self.border
            
            # Status
            status = record.get('simplified_status', record.get('current_status', 'N/A'))
            if 'CLOSED' in status or 'SECTION_COMPLETED' in status:
                status = 'Completed'
            status_cell = ws.cell(row=row_idx, column=5, value=status)
            status_cell.border = self.border
            
            # NOV
            has_nov = record.get('has_nov', False)
            nov_cell = ws.cell(row=row_idx, column=6, value='✓' if has_nov else '✗')
            nov_cell.border = self.border
            nov_cell.alignment = Alignment(horizontal='center', vertical='center')
            if has_nov:
                nov_cell.font = Font(name='Arial', size=12, color='00AA00', bold=True)
            else:
                nov_cell.font = Font(name='Arial', size=12, color='AA0000', bold=True)
            
            # NOO
            has_noo = record.get('has_noo', False)
            noo_cell = ws.cell(row=row_idx, column=7, value='✓' if has_noo else '✗')
            noo_cell.border = self.border
            noo_cell.alignment = Alignment(horizontal='center', vertical='center')
            if has_noo:
                noo_cell.font = Font(name='Arial', size=12, color='00AA00', bold=True)
            else:
                noo_cell.font = Font(name='Arial', size=12, color='AA0000', bold=True)
            
            # Compliance Status
            compliance = record.get('compliance_status', 'PENDING')
            compliance_cell = ws.cell(row=row_idx, column=8, value=compliance)
            compliance_cell.border = self.border
            if compliance == 'COMPLIANT':
                compliance_cell.fill = self.light_green_fill
            elif compliance == 'NON_COMPLIANT':
                compliance_cell.fill = self.light_red_fill
            
            # Inspected By
            ws.cell(row=row_idx, column=9, value=record.get('inspected_by_name', 'Not Inspected') or 'Not Inspected').border = self.border
        
        self._auto_adjust_column_width(ws)
    
    def _create_recommendations_sheet(self):
        """Create recommendations worksheet"""
        ws = self.workbook.create_sheet(title='Recommendations')
        
        # Title
        ws['A1'] = 'SYSTEM-GENERATED RECOMMENDATIONS'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')
        
        recommendations = self.report_data.get('recommendations', [])
        
        row = 3
        for idx, rec in enumerate(recommendations, start=1):
            ws[f'A{row}'] = f"{idx}. {rec.get('type', 'Recommendation')}"
            ws[f'A{row}'].font = self.bold_font
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws[f'A{row}'] = rec.get('description', '')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row}:C{row}')
            ws.row_dimensions[row].height = 40
            
            row += 2
        
        # Manual recommendations section
        row += 2
        ws[f'A{row}'] = 'MANUAL RECOMMENDATIONS'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].fill = self.light_blue_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = '(Space for division chief to add manual recommendations)'
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 100
        
        self._auto_adjust_column_width(ws)
    
    def generate(self):
        """Generate the complete Excel workbook"""
        self._create_summary_sheet()
        self._create_detailed_data_sheet()
        self._create_recommendations_sheet()
        
        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output

