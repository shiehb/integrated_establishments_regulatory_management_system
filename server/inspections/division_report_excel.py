"""
Division Report Excel Generator using openpyxl
Generates professional Excel reports with DENR official standards
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DivisionReportExcelGenerator:
    """
    Professional Excel generator for division reports
    """
    
    def __init__(self, report_data, filters_applied):
        self.report_data = report_data
        self.filters_applied = filters_applied
        self.workbook = Workbook()
        
        # Generate DENR reference number
        self.reference_number = self._generate_reference_number()
        
        # DENR colors
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')  # DENR Blue
        self.denr_green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # DENR Green
        self.subheader_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        self.light_blue_fill = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')
        self.light_green_fill = PatternFill(start_color='E7F7E7', end_color='E7F7E7', fill_type='solid')
        self.light_red_fill = PatternFill(start_color='FFE7E7', end_color='FFE7E7', fill_type='solid')
        
        # Define fonts - Arial as per DENR standards
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        # Define borders
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def _generate_reference_number(self):
        """Generate DENR reference number: EIA-YYYY-MM-DD-####"""
        now = datetime.now()
        date_str = now.strftime('%Y-%m-%d')
        sequence = str(int(now.timestamp() * 1000))[-4:]
        return f"EIA-{date_str}-{sequence}"
    
    def _add_denr_header(self, worksheet, start_row=1):
        """Add official DENR header to worksheet"""
        ws = worksheet
        row = start_row
        
        # DENR Header
        ws[f'A{row}'] = 'REPUBLIC OF THE PHILIPPINES'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='0066CC')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = 'DEPARTMENT OF ENVIRONMENT AND NATURAL RESOURCES'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='008000')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = 'ENVIRONMENTAL MANAGEMENT BUREAU'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='008000')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = 'REGION I'
        ws[f'A{row}'].font = Font(name='Arial', size=10, color='008000')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = 'Kalikasang Protektado, Paglilingkod na Tapat.'
        ws[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color='666666')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = f'Reference Number: {self.reference_number}'
        ws[f'A{row}'].font = self.bold_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        return row
    
    def _freeze_header_row(self, worksheet, header_row):
        """Freeze panes at header row"""
        worksheet.freeze_panes = worksheet[f'A{header_row + 1}']
    
    def _add_auto_filter(self, worksheet, start_row, end_row, num_cols):
        """Add auto-filter to data table"""
        start_col_letter = get_column_letter(1)
        end_col_letter = get_column_letter(num_cols)
        worksheet.auto_filter.ref = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        
    def _auto_adjust_column_width(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self):
        """Create summary statistics worksheet"""
        ws = self.workbook.active
        ws.title = 'Summary Statistics'
        
        # Add DENR header
        row = self._add_denr_header(ws)
        
        # Title
        ws[f'A{row}'] = 'DIVISION REPORT - SUMMARY STATISTICS'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Generation info
        ws[f'A{row}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws[f'A{row}'].font = self.normal_font
        row += 2
        
        # Legal bases section
        ws[f'A{row}'] = 'LEGAL BASES'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        legal_bases = [
            'RA 8749 - Clean Air Act',
            'RA 9275 - Clean Water Act',
            'RA 9003 - Ecological Solid Waste Management Act',
            'PD 1586 - EIS Law',
            'DAO 2016-08 (Procedural Manual for PEISS)',
            'DAO 1996-37 (Hazardous Waste)',
            'DAO 2021-19 (Updated Standards)',
            'EMB Memorandum Circulars and Regional Policies'
        ]
        row += 1
        for base in legal_bases:
            ws[f'A{row}'] = f'• {base}'
            ws[f'A{row}'].font = self.normal_font
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        # Filters applied
        row += 1
        ws[f'A{row}'] = 'FILTERS APPLIED'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for key, value in self.filters_applied.items():
            if value:
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = self.bold_font
                row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = 'INSPECTION SUMMARY'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws.merge_cells(f'A{row}:B{row}')
        
        stats = self.report_data.get('statistics', {})
        inspection_stats = stats.get('inspection_summary', {})
        compliance_stats = stats.get('compliance_summary', {})
        
        row += 1
        summary_data = [
            ('Total Inspections', inspection_stats.get('total_inspections', 0)),
            ('Division Reviewed', inspection_stats.get('division_reviewed', 0)),
            ('Section Completed', inspection_stats.get('section_completed', 0)),
            ('Total NOV Issued', inspection_stats.get('total_nov', 0)),
            ('Total NOO Issued', inspection_stats.get('total_noo', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.bold_font
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Compliance summary
        row += 2
        ws[f'A{row}'] = 'COMPLIANCE SUMMARY'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = self.denr_green_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        compliance_data = [
            ('Compliant', compliance_stats.get('compliant_count', 0)),
            ('Non-Compliant', compliance_stats.get('non_compliant_count', 0)),
            ('Pending', compliance_stats.get('pending_count', 0)),
        ]
        
        for label, value in compliance_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.bold_font
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        self._auto_adjust_column_width(ws)
    
    def _create_detailed_data_sheet(self):
        """Create detailed data worksheet"""
        ws = self.workbook.create_sheet(title='Detailed Data')
        
        # Add DENR header
        row = self._add_denr_header(ws)
        
        # Headers
        headers = [
            'Inspection No.', 'Establishment', 'Law', 'Inspection Date',
            'Status', 'NOV', 'NOO', 'Compliance Status', 'Inspected By'
        ]
        
        header_row = row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        records = self.report_data.get('records', [])
        row += 1
        data_start_row = row
        
        row_idx = row
        for record in records:
            # Inspection No.
            ws.cell(row=row_idx, column=1, value=record.get('code', 'N/A')).border = self.border
            
            # Establishment
            ws.cell(row=row_idx, column=2, value=record.get('establishment_name', 'N/A')).border = self.border
            
            # Law
            ws.cell(row=row_idx, column=3, value=record.get('law', 'N/A')).border = self.border
            
            # Inspection Date
            created_at = record.get('created_at', '')
            if created_at:
                created_at = created_at[:10]  # Extract date part
            ws.cell(row=row_idx, column=4, value=created_at or 'N/A').border = self.border
            
            # Status
            status = record.get('simplified_status', record.get('current_status', 'N/A'))
            if 'CLOSED' in status or 'SECTION_COMPLETED' in status:
                status = 'Completed'
            status_cell = ws.cell(row=row_idx, column=5, value=status)
            status_cell.border = self.border
            
            # NOV
            has_nov = record.get('has_nov', False)
            nov_cell = ws.cell(row=row_idx, column=6, value='✓' if has_nov else '✗')
            nov_cell.border = self.border
            nov_cell.alignment = Alignment(horizontal='center', vertical='center')
            if has_nov:
                nov_cell.font = Font(name='Arial', size=12, color='00AA00', bold=True)
            else:
                nov_cell.font = Font(name='Arial', size=12, color='AA0000', bold=True)
            
            # NOO
            has_noo = record.get('has_noo', False)
            noo_cell = ws.cell(row=row_idx, column=7, value='✓' if has_noo else '✗')
            noo_cell.border = self.border
            noo_cell.alignment = Alignment(horizontal='center', vertical='center')
            if has_noo:
                noo_cell.font = Font(name='Arial', size=12, color='00AA00', bold=True)
            else:
                noo_cell.font = Font(name='Arial', size=12, color='AA0000', bold=True)
            
            # Compliance Status
            compliance = record.get('compliance_status', 'PENDING')
            compliance_cell = ws.cell(row=row_idx, column=8, value=compliance)
            compliance_cell.border = self.border
            if compliance == 'COMPLIANT':
                compliance_cell.fill = self.light_green_fill
            elif compliance == 'NON_COMPLIANT':
                compliance_cell.fill = self.light_red_fill
            
            # Inspected By
            ws.cell(row=row_idx, column=9, value=record.get('inspected_by_name', 'Not Inspected') or 'Not Inspected').border = self.border
            
            # Apply alternating row colors
            if row_idx % 2 == 0:
                for col in range(1, 10):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.fill.start_color.index == '00000000':  # Only if no fill already
                        cell.fill = self.light_blue_fill
            
            row_idx += 1
        
        data_end_row = row_idx - 1 if records else header_row
        
        # Add auto-filter
        if data_end_row > header_row:
            self._add_auto_filter(ws, header_row, data_end_row, len(headers))
        
        # Freeze header row
        self._freeze_header_row(ws, header_row)
        
        # Add routing section
        self._add_routing_section(ws, data_end_row + 2)
        
        self._auto_adjust_column_width(ws)
    
    def _add_routing_section(self, worksheet, start_row):
        """Add DENR routing section for workflow tracking"""
        ws = worksheet
        row = start_row
        
        ws[f'A{row}'] = 'ROUTING AND APPROVAL'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:I{row}')
        row += 1
        
        routing_headers = ['Stage', 'Name', 'Position', 'Date', 'Signature']
        for col, header in enumerate(routing_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        row += 1
        routing_data = [
            ['Prepared by', '', 'Monitoring Staff', '', ''],
            ['Reviewed by', '', 'Section Chief', '', ''],
            ['Recommended by', '', 'Division Chief', '', ''],
            ['Approved by', '', 'Regional Director', '', ''],
        ]
        
        for routing_row_data in routing_data:
            for col, value in enumerate(routing_row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row % 2 == 0:
                    cell.fill = self.light_blue_fill
            row += 1
    
    def _create_recommendations_sheet(self):
        """Create recommendations worksheet"""
        ws = self.workbook.create_sheet(title='Recommendations')
        
        # Add DENR header
        row = self._add_denr_header(ws)
        
        # Title
        ws[f'A{row}'] = 'SYSTEM-GENERATED RECOMMENDATIONS'
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        recommendations = self.report_data.get('recommendations', [])
        
        row += 1
        for idx, rec in enumerate(recommendations, start=1):
            ws[f'A{row}'] = f"{idx}. {rec.get('type', 'Recommendation')}"
            ws[f'A{row}'].font = self.bold_font
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws[f'A{row}'] = rec.get('description', '')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row}:C{row}')
            ws.row_dimensions[row].height = 40
            
            row += 2
        
        # Manual recommendations section
        row += 2
        ws[f'A{row}'] = 'MANUAL RECOMMENDATIONS'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].fill = self.light_blue_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = '(Space for division chief to add manual recommendations)'
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 100
        
        self._auto_adjust_column_width(ws)
    
    def generate(self):
        """Generate the complete Excel workbook"""
        self._create_summary_sheet()
        self._create_detailed_data_sheet()
        self._create_recommendations_sheet()
        
        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output

