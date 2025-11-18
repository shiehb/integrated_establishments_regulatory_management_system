"""
Legal Report Excel Generator using openpyxl
Generates professional Excel reports with multiple worksheets
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class LegalReportExcelGenerator:
    """
    Professional Excel generator for legal reports
    """
    
    def __init__(self, report_data, filters_applied):
        self.report_data = report_data
        self.filters_applied = filters_applied
        self.workbook = Workbook()
        
        # Define colors
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        self.light_blue_fill = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')
        self.light_green_fill = PatternFill(start_color='E7F7E7', end_color='E7F7E7', fill_type='solid')
        self.light_red_fill = PatternFill(start_color='FFE7E7', end_color='FFE7E7', fill_type='solid')
        
        # Define fonts
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        # Define borders
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
    def _auto_adjust_column_width(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self):
        """Create summary statistics worksheet"""
        ws = self.workbook.active
        ws.title = 'Summary Statistics'
        
        # Title
        ws['A1'] = 'LEGAL REPORT - SUMMARY STATISTICS'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        
        # Generation info
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = self.normal_font
        
        # Filters applied
        row = 4
        ws[f'A{row}'] = 'FILTERS APPLIED'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for key, value in self.filters_applied.items():
            if value:
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = self.bold_font
                row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = 'BILLING SUMMARY'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.light_blue_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        stats = self.report_data.get('statistics', {})
        billing_stats = stats.get('billing_summary', {})
        
        row += 1
        summary_data = [
            ('Total Billed Amount', f"₱{billing_stats.get('total_billed', 0):,.2f}"),
            ('Total Paid Amount', f"₱{billing_stats.get('total_paid', 0):,.2f}"),
            ('Outstanding Balance', f"₱{billing_stats.get('outstanding_balance', 0):,.2f}"),
            ('Average Days to Payment', f"{billing_stats.get('avg_days_to_payment', 0):.1f} days"),
            ('Total NOV Issued', billing_stats.get('total_nov', 0)),
            ('Total NOO Issued', billing_stats.get('total_noo', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.bold_font
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Compliance summary
        row += 2
        ws[f'A{row}'] = 'COMPLIANCE SUMMARY'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = self.light_green_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        compliance_stats = stats.get('compliance_summary', {})
        
        row += 1
        compliance_data = [
            ('Compliant Establishments', compliance_stats.get('compliant_count', 0)),
            ('Non-Compliant Establishments', compliance_stats.get('non_compliant_count', 0)),
            ('Pending Actions', compliance_stats.get('pending_count', 0)),
            ('Re-inspections Recommended', compliance_stats.get('reinspection_recommended', 0)),
        ]
        
        for label, value in compliance_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.bold_font
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        self._auto_adjust_column_width(ws)
    
    def _create_detailed_data_sheet(self):
        """Create detailed data worksheet"""
        ws = self.workbook.create_sheet(title='Detailed Data')
        
        # Headers
        headers = [
            'Inspection No.', 'Establishment', 'Billing Amount', 'Billing Date',
            'Payment Status', 'Payment Date', 'NOV/NOO', 'Compliance Status',
            'Legal Actions', 'Remarks', 'Assigned Legal Officer'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        records = self.report_data.get('records', [])
        for row_idx, record in enumerate(records, start=2):
            # Inspection No.
            ws.cell(row=row_idx, column=1, value=record.get('inspection_code', 'N/A')).border = self.border
            
            # Establishment
            ws.cell(row=row_idx, column=2, value=record.get('establishment_name', 'N/A')).border = self.border
            
            # Billing Amount
            amount_cell = ws.cell(row=row_idx, column=3, value=float(record.get('amount', 0)))
            amount_cell.number_format = '₱#,##0.00'
            amount_cell.border = self.border
            
            # Billing Date
            billing_date = record.get('sent_date', '')
            if billing_date:
                billing_date = datetime.fromisoformat(billing_date.replace('Z', '+00:00')).strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=4, value=billing_date).border = self.border
            
            # Payment Status
            payment_status = record.get('payment_status', 'UNPAID')
            status_cell = ws.cell(row=row_idx, column=5, value=payment_status)
            status_cell.border = self.border
            if payment_status == 'PAID':
                status_cell.fill = self.light_green_fill
            else:
                status_cell.fill = self.light_red_fill
            
            # Payment Date
            payment_date = record.get('payment_date', '')
            ws.cell(row=row_idx, column=6, value=payment_date or 'N/A').border = self.border
            
            # NOV/NOO
            nov_noo = []
            if record.get('has_nov'):
                nov_noo.append('NOV')
            if record.get('has_noo'):
                nov_noo.append('NOO')
            ws.cell(row=row_idx, column=7, value=', '.join(nov_noo) if nov_noo else 'None').border = self.border
            
            # Compliance Status
            compliance = record.get('compliance_status', 'PENDING')
            compliance_cell = ws.cell(row=row_idx, column=8, value=compliance)
            compliance_cell.border = self.border
            if compliance == 'COMPLIANT':
                compliance_cell.fill = self.light_green_fill
            elif compliance == 'NON_COMPLIANT':
                compliance_cell.fill = self.light_red_fill
            
            # Legal Actions
            legal_action = record.get('legal_action', 'NONE')
            ws.cell(row=row_idx, column=9, value=legal_action.replace('_', ' ').title()).border = self.border
            
            # Remarks
            remarks = record.get('payment_notes', '') or record.get('recommendations', '')
            ws.cell(row=row_idx, column=10, value=remarks[:100] if remarks else 'N/A').border = self.border
            
            # Assigned Legal Officer
            ws.cell(row=row_idx, column=11, value=record.get('assigned_legal_officer', 'N/A')).border = self.border
        
        # Add totals row
        if records:
            totals_row = len(records) + 2
            ws.cell(row=totals_row, column=1, value='TOTAL').font = self.bold_font
            ws.cell(row=totals_row, column=1).fill = self.subheader_fill
            
            # Calculate total amount
            total_formula = f'=SUM(C2:C{len(records) + 1})'
            total_cell = ws.cell(row=totals_row, column=3, value=total_formula)
            total_cell.font = self.bold_font
            total_cell.fill = self.subheader_fill
            total_cell.number_format = '₱#,##0.00'
        
        self._auto_adjust_column_width(ws)
    
    def _create_recommendations_sheet(self):
        """Create recommendations worksheet"""
        ws = self.workbook.create_sheet(title='Recommendations')
        
        # Title
        ws['A1'] = 'SYSTEM-GENERATED RECOMMENDATIONS'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')
        
        recommendations = self.report_data.get('recommendations', [])
        
        row = 3
        for idx, rec in enumerate(recommendations, start=1):
            ws[f'A{row}'] = f"{idx}. {rec.get('type', 'Recommendation')}"
            ws[f'A{row}'].font = self.bold_font
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws[f'A{row}'] = rec.get('description', '')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row}:C{row}')
            ws.row_dimensions[row].height = 40
            
            row += 2
        
        # Manual recommendations section
        row += 2
        ws[f'A{row}'] = 'MANUAL RECOMMENDATIONS'
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].fill = self.light_blue_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = '(Space for legal officer to add manual recommendations)'
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 100
        
        self._auto_adjust_column_width(ws)
    
    def generate(self):
        """Generate the complete Excel workbook"""
        self._create_summary_sheet()
        self._create_detailed_data_sheet()
        self._create_recommendations_sheet()
        
        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output

